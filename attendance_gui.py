import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl

class AttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Management System")
        self.root.geometry("1200x800")  # Set window size for laptops
        self.root.configure(bg="white")  # Background color

        self.excel_file = None

        # Main Page Layout
        self.create_main_page()

    def create_main_page(self):
        # Center Title
        self.title_label = tk.Label(
            self.root,
            text="Ahmedabad University",
            font=("Helvetica", 24, "bold"),
            bg="white",
        )
        self.title_label.pack(pady=10)

        # Sub Title
        self.subtitle_label = tk.Label(
            self.root,
            text="Attendance Management System",
            font=("Helvetica", 18, "bold"),
            bg="white",
        )
        self.subtitle_label.pack(pady=10)

        # Frame for buttons
        button_frame = tk.Frame(self.root, bg="white")
        button_frame.pack(pady=20)

        # File label
        self.file_label = tk.Label(
            button_frame, text="No file selected", width=50, anchor="w", bg="white"
        )
        self.file_label.grid(row=0, column=0, padx=10, pady=10)

        # Browse Button
        self.browse_button = tk.Button(
            button_frame,
            text="Browse Excel File",
            command=self.browse_file,
            bg="#FFCCCC",
            font=("Helvetica", 14),
        )
        self.browse_button.grid(row=0, column=1, padx=10, pady=10)

        # Update Button
        self.update_button = tk.Button(
            button_frame,
            text="Update Attendance",
            command=self.open_update_window,
            bg="#FF9999",
            font=("Helvetica", 14),
            state="disabled",
        )
        self.update_button.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

        # View Button
        self.view_button = tk.Button(
            button_frame,
            text="View Attendance",
            command=self.open_view_window,
            bg="#FF6666",
            font=("Helvetica", 14),
            state="disabled",
        )
        self.view_button.grid(row=2, column=0, columnspan=2, padx=10, pady=10)

        # Add Button
        self.add_button = tk.Button(
            button_frame,
            text="Add Attendance",
            command=self.open_add_window,
            bg="#FFCCCC",
            font=("Helvetica", 14),
            state="disabled",
        )
        self.add_button.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

    def browse_file(self):
        self.excel_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if self.excel_file:
            self.file_label.config(text=self.excel_file)
            self.update_button.config(state="normal")
            self.view_button.config(state="normal")
            self.add_button.config(state="normal")

    def open_update_window(self):
        if not self.excel_file:
            messagebox.showerror("Error", "No file selected.")
            return

        UpdateAttendanceWindow(self.root, self.excel_file)

    def open_view_window(self):
        if not self.excel_file:
            messagebox.showerror("Error", "No file selected.")
            return

        ViewAttendanceWindow(self.root, self.excel_file)

    def open_add_window(self):
        if not self.excel_file:
            messagebox.showerror("Error", "No file selected.")
            return

        UpdateAttendanceWindow(self.root, self.excel_file)

    def open_view_window(self):
        if not self.excel_file:
            messagebox.showerror("Error", "No file selected.")
            return

        ViewAttendanceWindow(self.root, self.excel_file)

    def open_add_window(self):
        if not self.excel_file:
            messagebox.showerror("Error", "No file selected.")
            return

        AddAttendanceWindow(self.root, self.excel_file)


class UpdateAttendanceWindow:
    def __init__(self, parent, excel_file):
        self.excel_file = excel_file
        self.window = tk.Toplevel(parent)
        self.window.title("Update Attendance")

        # Labels and entries for roll number, day, and attendance
        tk.Label(self.window, text="Roll Number:").grid(row=0, column=0, padx=10, pady=10)
        self.roll_entry = tk.Entry(self.window)
        self.roll_entry.grid(row=0, column=1, padx=10, pady=10)

        tk.Label(self.window, text="Day:").grid(row=1, column=0, padx=10, pady=10)
        self.day_entry = tk.Entry(self.window)
        self.day_entry.grid(row=1, column=1, padx=10, pady=10)

        tk.Label(self.window, text="Attendance (1=Present, 0=Absent):").grid(row=2, column=0, padx=10, pady=10)
        self.attendance_entry = tk.Entry(self.window)
        self.attendance_entry.grid(row=2, column=1, padx=10, pady=10)

        # Update button
        self.update_button = tk.Button(self.window, text="Update", command=self.update_attendance)
        self.update_button.grid(row=3, column=0, columnspan=2, pady=10)

    def update_attendance(self):
        roll_number = self.roll_entry.get().strip()
        day = self.day_entry.get().strip()
        attendance = self.attendance_entry.get().strip()

        if not roll_number or not day or not attendance:
            messagebox.showerror("Error", "Please fill in all fields.")
            return

        if not roll_number.isdigit() or not (2340001 <= int(roll_number) <= 2340148):
            messagebox.showerror("Error", "Roll number must be between 2340001 and 2340148.")
            return

        if not day.isdigit() or not (1 <= int(day) <= 31):
            messagebox.showerror("Error", "Day must be between 1 and 31.")
            return

        if attendance not in ("0", "1"):
            messagebox.showerror("Error", "Attendance must be 1 (Present) or 0 (Absent).")
            return

        try:
            day = int(day)
            attendance = int(attendance)
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active

            # Find roll number in Excel sheet
            roll_number_found = False
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
                if str(row[0].value) == roll_number:
                    row_idx = row[0].row
                    sheet.cell(row=row_idx, column=day + 1, value=attendance)
                    wb.save(self.excel_file)
                    wb.close()
                    messagebox.showinfo("Success", "Attendance updated successfully.")
                    roll_number_found = True
                    break

            if not roll_number_found:
                messagebox.showerror("Error", "Roll number not found.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update attendance: {e}")


class ViewAttendanceWindow:
    def __init__(self, parent, excel_file):
        self.excel_file = excel_file
        self.window = tk.Toplevel(parent)
        self.window.title("View Attendance")

        self.tree = ttk.Treeview(self.window, columns=("Roll Number", "Present", "Absent", "Percentage"), show="headings", height=20)
        self.tree.heading("Roll Number", text="Roll Number")
        self.tree.heading("Present", text="Present Days")
        self.tree.heading("Absent", text="Absent Days")
        self.tree.heading("Percentage", text="Percentage (%)")
        self.tree.column("Roll Number", width=150, anchor="center")
        self.tree.column("Present", width=100, anchor="center")
        self.tree.column("Absent", width=100, anchor="center")
        self.tree.column("Percentage", width=150, anchor="center")
        self.tree.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        self.scrollbar = ttk.Scrollbar(self.window, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.grid(row=0, column=1, sticky="ns")

        self.load_attendance()

    def load_attendance(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                roll_number = row[0]
                present_days = sum(1 for val in row[1:] if val == 1)
                total_days = len(row) - 1
                absent_days = total_days - present_days
                percentage = (present_days / total_days) * 100 if total_days > 0 else 0

                # Insert row into Treeview with color coding
                item = self.tree.insert("", "end", values=(roll_number, present_days, absent_days, f"{percentage:.2f}"))
                if percentage < 80:
                    self.tree.item(item, tags=("low",))
                elif 95 <= percentage < 100:
                    self.tree.item(item, tags=("high",))
                elif percentage == 100:
                    self.tree.item(item, tags=("perfect",))

            # Style Tags
            self.tree.tag_configure("low", background="red", foreground="white")
            self.tree.tag_configure("high", background="green", foreground="white")
            self.tree.tag_configure("perfect", background="gold", foreground="black")

            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load attendance: {e}")


class AddAttendanceWindow:
    def __init__(self, parent, excel_file):
        self.excel_file = excel_file
        self.window = tk.Toplevel(parent)
        self.window.title("Add Attendance")

        # Load Excel file
        try:
            self.wb = openpyxl.load_workbook(self.excel_file)
            self.sheet = self.wb.active

            # Determine the next day column
            self.next_day = self.sheet.max_column - 1 + 1  # Skip Roll Number column
            self.day_label_text = f"Day {self.next_day}"

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {e}")
            return

        # Header
        self.header_label = tk.Label(self.window, text=f"Add Attendance for {self.day_label_text}", font=("Arial", 16))
        self.header_label.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

        # Scrollable Frame for Roll Numbers and Checkboxes
        self.canvas = tk.Canvas(self.window)
        self.scrollbar = ttk.Scrollbar(self.window, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=1, column=1, sticky="ns", padx=10, pady=10)

        # Dynamically create roll number and checkbox entries
        self.attendance_data = {}  # Dictionary to store roll numbers and attendance
        row_idx = 0

        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=True):
            roll_number = row[0]

            roll_label = tk.Label(self.scrollable_frame, text=f"Roll Number: {roll_number}", anchor="w")
            roll_label.grid(row=row_idx, column=0, padx=10, pady=5)

            var = tk.IntVar(value=0)  # 0 = Absent, 1 = Present
            checkbox = tk.Checkbutton(self.scrollable_frame, text="Present", variable=var)
            checkbox.grid(row=row_idx, column=1, padx=10, pady=5)

            # Store the IntVar reference in the attendance_data
            self.attendance_data[roll_number] = var
            row_idx += 1

        # Save Button
        self.save_button = tk.Button(self.window, text="Save Attendance", command=self.save_attendance)
        self.save_button.grid(row=2, column=0, columnspan=2, pady=10)

    def save_attendance(self):
        try:
            # Add a new column for the next day
            self.sheet.cell(row=1, column=self.next_day + 1, value=f"Day {self.next_day}")

            for i, (roll_number, var) in enumerate(self.attendance_data.items(), start=2):
                attendance = var.get()  # 0 for Absent, 1 for Present
                self.sheet.cell(row=i, column=self.next_day + 1, value=attendance)

            # Save the workbook
            self.wb.save(self.excel_file)
            self.wb.close()
            messagebox.showinfo("Success", "Attendance added successfully.")
            self.window.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save attendance: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceApp(root)
    root.mainloop()
